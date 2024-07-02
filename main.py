import os
import pandas as pd
import openpyxl
from tkinter import filedialog as fd


def syb(x: int, y: int, a: int):
    if sheet.cell(x, y).value == None:
        sheet.cell(x, y).value = f"={a}"
    else:
        sheet.cell(x, y).value += f"+{a}"


def dr(d):
    dat = pd.Timestamp(d.iat[r, 0])
    amount = d.iat[r, 4]
    dy = dat.day
    match d.iat[r, 3]:
        case "Cash":
            syb(dy, 2, amount)
        case "O-Card":
            syb(dy, 4, amount)
        case "Saving_HSBC":
            syb(dy, 6, amount)
        case "O-Card Samsung Pay":
            syb(dy, 6, amount)


def cr(d):
    dat = pd.Timestamp(d.iat[r, 0])
    amount = d.iat[r, 4]
    dy = dat.day
    match d.iat[r, 3]:
        case "Cash":
            syb(dy, 3, amount)
        case "O-Card":
            syb(dy, 5, amount)
        case "Saving_HSBC":
            syb(dy, 7, amount)
        case "O-Card Samsung Pay":
            syb(dy, 7, amount)


def path():
    path: str = fd.askopenfilename()
    return path

list = []
file = path()
month = int(input("Month: "))
r = 0
df = pd.read_excel(file)
df = df.iloc[:, 0:12]
wb = openpyxl.Workbook()
sheet = wb.worksheets[0]
while r < len(df):
    date = pd.Timestamp(df.iat[r, 0])
    if date.month != month:
        list.append(r)
        print(f"Wrong month in line {r}")
        r += 1
        continue
    dy = date.day
    amount = df.iat[r, 4]
    match df.iat[r, 6]:
        case "INCOME":
            sheet.cell(date.day, 1).value = df.iat[r, 4]
            dr(df)
        case "EXPENSE":
            cr(df)
            match df.iat[r, 2]:
                case "Shopping":
                    syb(dy, 8, amount)
                case "Food & Drinks":
                    syb(dy, 9, amount)
                case "Transport":
                    syb(dy, 10, amount)
                case "Bills & Fees":
                    syb(dy, 11, amount)
                case "Entertainment":
                    syb(dy, 12, amount)
        case "TRANSFER":
            cr(df)
            match df.iat[r, 9]:
                case "Cash":
                    syb(dy, 2, amount)
                case "O-Card":
                    syb(dy, 4, amount)
                case "Saving_HSBC":
                    syb(dy, 6, amount)
                case "O-Card Samsung Pay":
                    syb(dy, 6, amount)
    r += 1
wb.save(r"C:\Desktop\test.xlsx")
print("Done")
print(f"Error list: {list}")
input("Press Enter to continue...")
